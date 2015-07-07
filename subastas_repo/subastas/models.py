# -*- coding: utf-8 -*-

from django.db import IntegrityError
from django.db import models
from django.db import transaction
from django.db.models.query import QuerySet
from django.utils import timezone

from openpyxl import load_workbook


class Lote(models.Model):
    grupo = models.ForeignKey('Grupo',
                              related_name='lotes',
                              blank=True, null=True)
    numero = models.IntegerField()
    subastado = models.BooleanField(default=False)
    chatarra = models.BooleanField(default=False)

    def __unicode__(self):
        return "%s" % self.numero

    class Meta:
        unique_together = ('grupo', 'numero')


class Tipo(models.Model):
    nombre = models.CharField(max_length=100)

    def __unicode__(self):
        return self.nombre


class RodadoQuerySet(QuerySet):
    def load_bienes(self, path_xlsx):
        wb = load_workbook(path_xlsx)
        ws = wb.active
        instances = []
        for row in ws.iter_rows(range_string="A2:K4"):
            try:
                with transaction.atomic():
                    rodado = Rodado.objects \
                        .create(numero_inventario=row[0].value,
                                descripcion=row[4].value,
                                modelo=row[8].value,
                                chasis=row[9].value,
                                motor=row[10].value,
                                dominio=row[6].value)
            except IntegrityError, e:
                print e
                continue
            else:
                instances.append(rodado)
        return len(instances)

    def no_subastados(self):
        return self.filter(subastado=False)

    def subastados(self):
        return self.filter(subastado=True)


class Rodado(models.Model):
    lote = models.ForeignKey(Lote,
                             related_name="bienes",
                             blank=True, null=True)
    tipo = models.ForeignKey(Tipo)
    numero_inventario = models.IntegerField(unique=True)
    descripcion = models.TextField(blank=True, null=True)
    modelo = models.CharField(max_length=50)
    chasis = models.CharField(max_length=50)
    motor = models.CharField(max_length=50)
    dominio = models.CharField(max_length=50)
    marca = models.CharField(max_length=100)
    anio = models.IntegerField("Año", blank=True, null=True)
    precio_base = models.FloatField(default=0)
    precio_venta = models.FloatField(default=0)

    subastado = models.BooleanField(default=False)

    objects = RodadoQuerySet.as_manager()

    def __unicode__(self):
        return "%s %s %s" % (self.chasis, self.motor, self.dominio)


class SubastaManager(models.Manager):
    def get_current(self):
        return super(SubastaManager, self).get_queryset() \
            .filter(cerrado_el=None).last()


class Subasta(models.Model):
    numero = models.IntegerField()
    fecha_hora = models.DateTimeField()
    cerrado_el = models.DateTimeField(blank=True, null=True)
    decreto = models.CharField(max_length=10)
    domicilio = models.ForeignKey('personas.Domicilio')

    profesionales = models.ManyToManyField('personas.Profesional')
    personas = models.ManyToManyField('personas.Persona',
                                      blank=True, null=True)
    created_at = models.DateTimeField(auto_now=True)
    updated_at = models.DateTimeField(auto_now_add=True)
    user_updated = models.ForeignKey('users.User', blank=True, null=True)

    objects = SubastaManager()

    def __unicode__(self):
        return "%s" % self.fecha_hora

    def close(self):
        self.cerrado_el = timezone.now()
        self.save()

    @property
    def lotes(self):
        grupos = self.grupos.all()
        return Lote.objects.filter(grupo__in=grupos)


class Grupo(models.Model):
    subasta = models.ForeignKey(Subasta,
                                related_name='grupos',
                                blank=True, null=True)
    numero = models.IntegerField()
    subastado = models.BooleanField(default=False)
    martillero = models.ForeignKey('personas.Profesional',
                                   blank=True,
                                   null=True)

    def __unicode__(self):
        return "%s" % self.numero

    class Meta:
        unique_together = ('subasta', 'numero')


class Acta(models.Model):
    subasta = models.ForeignKey('subastas.Subasta', related_name='actas')
    lote = models.OneToOneField(Lote)
    persona = models.ForeignKey('personas.Persona')
    profesionales = models.ManyToManyField('personas.Profesional')
    descripcion = models.TextField(blank=True, null=True)

    def __unicode__(self):
        return "Lote: %s comprado por: %s" % (self.lote, self.persona)

    # def save(self, *args, **kwargs):
    #     for bien in Rodado.objects.filter(lote=self.lote):
    #         bien.subastado = True
    #         bien.save()
    #     super(Acta, self).save(*args, **kwargs)
